import os, glob, time, torch, sys
import numpy as np
from PIL import Image
from PySide6.QtCore import Qt
from PySide6.QtGui import QPixmap, QFont, QFontDatabase
from PySide6.QtWidgets import QApplication, QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout, QFileDialog, QTableWidget, QTableWidgetItem
from torchvision import transforms
from importlib.resources import files
from MycoRadar.unet import UNet
import openpyxl
from openpyxl.styles import Font
DEVICE = torch.device("mps" if torch.backends.mps.is_available() else "cpu")

model = UNet(in_channels=3, out_channels=3)
model_path = files("MycoRadar.Models").joinpath("UNET.pth")
model.load_state_dict(torch.load(model_path, map_location=DEVICE))
model.to(DEVICE)
model.eval()

class UI(QWidget):
    def __init__(self):
        super().__init__()
        self.folder_path = ""
        self.i = 0
        self.image_paths = sorted(glob.glob(os.path.join(self.folder_path, "*.jpg")))
        self.setWindowTitle("MycoRadar")
        self.setStyleSheet("""
            QWidget {
                background-color: #ffffff;
                color: #333333;
            }
            QLabel {
                fon3t-size: 20px;
                font-weight: 600;
                color: #2f5d2f;
                margin-bottom: 12px;
            }
            QPushButton {
                background-color: #2e7d32;
                color: white;
                border: none;
                padding: 10px 20px;
                font-size: 14px;
                border-radius: 8px;
                font-weight: 600;
            }
            QPushButton:hover {
                background-color: #27642b;
            }
            QPushButton:pressed {
                background-color: #1b3d1b;
            }
        """)

        mainLayout = QHBoxLayout()
        subLayout = QVBoxLayout()
        sub2layout = QVBoxLayout()
        sub3layout = QHBoxLayout()
        subLayout.setAlignment(Qt.AlignmentFlag.AlignTop)
        sub2layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        self.title = QLabel("MycoRadar")
        self.title2 = QLabel()
        self.title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.title2.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subLayout.addWidget(self.title)

        self.image = QLabel()
        subLayout.addWidget(self.image)

        check_btn = QPushButton("Check")
        check_btn.clicked.connect(self.check)
        subLayout.addWidget(check_btn)

        smallLayout = QHBoxLayout()
        back_btn = QPushButton("Back")
        back_btn.clicked.connect(self.back_image)
        smallLayout.addWidget(back_btn)
        next_btn = QPushButton("Next")
        next_btn.clicked.connect(self.next_image)
        smallLayout.addWidget(next_btn)
        subLayout.addLayout(smallLayout)

        training = QPushButton("Choose Folder")
        training.clicked.connect(self.choose_folder)
        subLayout.addWidget(training)
        mainLayout.addLayout(subLayout)

        self.res = QLabel()
        sub2layout.addWidget(self.title2)
        sub2layout.addWidget(self.res)

        self.stats = QLabel(" ")
        self.stats2 = QLabel(" ")
        self.stats3 = QLabel(" ")
        self.check_all = QPushButton("Check All")
        self.check_all.clicked.connect(self.check_all_images)

        self.stats.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.res.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.stats2.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.stats3.setAlignment(Qt.AlignmentFlag.AlignCenter)

        sub3layout.addWidget(self.stats)
        sub3layout.addWidget(self.stats2)
        sub3layout.addWidget(self.stats3)
        sub2layout.addWidget(self.check_all)
        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        sub2layout.addWidget(self.export_btn)
        sub2layout.addLayout(sub3layout)
        mainLayout.addLayout(sub2layout)

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Image", "Root px", "AM px", "Colonized %"])
        self.table.setMinimumWidth(450)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setColumnWidth(0, 220)
        self.table.setColumnWidth(1, 75)
        self.table.setColumnWidth(2, 75)
        self.table.setColumnWidth(3, 85)
        tableLayout = QVBoxLayout()
        tableLayout.addWidget(self.table)
        mainLayout.addLayout(tableLayout)

        self.set_image(str(files("MycoRadar.Assets").joinpath("placeholder.png")))
        self.set_heatmap(str(files("MycoRadar.Assets").joinpath("placeholder.png")))
        self.setLayout(mainLayout)

    def set_image(self, original_image_path):
        self.image.clear()
        self.image.setPixmap(QPixmap(original_image_path).scaled(600, 450, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))

    def set_heatmap(self, original_image_path):
        self.res.clear()
        self.res.setPixmap(QPixmap(original_image_path).scaled(600, 450, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))

    def export_to_excel(self):
        if self.table.rowCount() == 0:
            self.title2.setText("No data to export")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "MycoRadar.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MycoRadar Results"


        headers = ["Image", "Root px", "AM px", "Colonized %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)


        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                value = self.table.item(row, col).text()
                ws.cell(row=row + 2, column=col + 1, value=value)

        try:
            wb.save(file_path)
            self.title2.setText(f"Exported to: {os.path.basename(file_path)}")
        except Exception as e:
            self.title2.setText(f"Export failed: {str(e)}")

    def check(self):
        if self.i >= len(self.image_paths):
            return

        transform = transforms.Compose([
            transforms.Resize((450, 600)),
            transforms.ToTensor()
        ])
        im = Image.open(self.image_paths[self.i]).convert("RGB")
        inp = transform(im).unsqueeze(0).to(DEVICE)
        start_time = time.time()
        with torch.no_grad():
            output = model(inp)
            softmax = torch.softmax(output, dim=1)
            pred = torch.argmax(softmax, dim=1).squeeze().cpu().numpy()

        overlay = np.zeros((pred.shape[0], pred.shape[1], 3), dtype=np.uint8)
        overlay[pred == 0] = [30, 60, 180]
        overlay[pred == 1] = [100, 220, 100]
        overlay[pred == 2] = [240, 50, 30]

        root_pixels = np.sum(pred == 1)
        am_pixels = np.sum(pred == 2)
        percent_colonized = (100 * (am_pixels / (am_pixels + root_pixels))) if root_pixels > 0 else 0

        self.stats.setText("# of root px's \n" + str(root_pixels))
        self.stats2.setText("# of am px's \n" + str(am_pixels))
        self.stats3.setText("Colonized \n" + "{:,.2f}".format(percent_colonized) + "%")
        filename = os.path.basename(self.image_paths[self.i])

        already_logged = False
        for row in range(self.table.rowCount()):
            if self.table.item(row, 0).text() == filename:
                already_logged = True
                break

        if not already_logged:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(filename))
            self.table.setItem(row, 1, QTableWidgetItem(str(root_pixels)))
            self.table.setItem(row, 2, QTableWidgetItem(str(am_pixels)))
            self.table.setItem(row, 3, QTableWidgetItem(f"{percent_colonized:.2f}%"))

        Image.fromarray(overlay).save("result.jpg")
        self.set_heatmap("result.jpg")
        os.remove("result.jpg")
        elapsed = time.time() - start_time
        self.title2.setText(f"Done {elapsed:.2f} sec")
        QApplication.processEvents()

    def next_image(self):
        self.i += 1
        if self.i < len(self.image_paths):
            self.set_image(self.image_paths[self.i])
        else:
            self.i = len(self.image_paths) - 1
            self.title2.setText("No more images")

    def back_image(self):
        if self.i > 0:
            self.i -= 1
            self.set_image(self.image_paths[self.i])
        else:
            self.i = 0
            self.title2.setText("Already at beginning")

    def choose_folder(self):
        self.folder_path = QFileDialog.getExistingDirectory(self, "Select Image Folder")
        self.image_paths = sorted(glob.glob(os.path.join(self.folder_path, "*.jpg")))
        self.i = 0
        if self.image_paths:
            self.set_image(self.image_paths[self.i])
        else:
            self.title2.setText("No images in this folder")

    def check_all_images(self):
        l = len(self.image_paths) - self.i
        for _ in range(l):
            self.title.setText( f" {_+1}/{l}  {(((_+1)/l)*100):.2f}%")
            self.check()
            self.next_image()

def main():
    app = QApplication(sys.argv)
    font_path = files("MycoRadar.Assets").joinpath("Montserrat-VariableFont_wght.ttf")
    font_id = QFontDatabase.addApplicationFont(str(font_path))
    families = QFontDatabase.applicationFontFamilies(font_id)
    if families:
        app.setFont(QFont(families[0]))
    window = UI()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
